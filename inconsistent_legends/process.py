# -*- coding: utf-8 -*-
import csv
import os
import random
import pandas as pd
import requests
type = ["Data-to-Data", "Data-to-VIS", "VIS-to-User", "User-to-VIS", "VIS-to-Data"]
rate = [[223, 898 - 223], [386, 898 - 386], [215, 898 - 215], [432, 898 - 432], [245, 898 - 245]]

def random_index(rate):
    # """随机变量的概率函数"""
    # 参数rate为list<int>
    # 返回概率事件的下标索引
    start = 0
    index = 0
    randnum = random.randint(1, sum(rate))
    for index, scope in enumerate(rate):
        start += scope
        if randnum <= start:
            break
    return index

def generate_labels():
    file_path = "./data.csv"
    csv_data = pd.read_csv(file_path, encoding='UTF-8', low_memory=False)  # 防止弹出警告
    df = pd.DataFrame(csv_data)
    labels=df['labels']
    for cnt in range(0, len(labels)):
        this_label = []
        for i in range(0, len(type)):
            if random_index(rate[i]):
                this_label.append(type[i])
        if len(this_label) == 0:
            this_label = ""
        else:
            this_label = "、".join(this_label)
        df['labels'][cnt] = this_label
    df.to_csv(file_path, index=False, encoding='utf_8_sig')

def download_img(img_url, api_token):
    print (img_url)
    header = {"Authorization": "Bearer " + api_token} # 设置http header，视情况加需要的条目，这里的token是用来鉴权的一种方式
    r = requests.get(img_url, headers=header, stream=True)
    if r.status_code == 200:
        img_path = './img/' + img_url.replace('http://vis.pku.edu.cn/', '').replace('/', '__')
        open(img_path, 'wb').write(r.content) # 将内容写入图片
        print("done")
    del r

basedir = os.path.abspath(os.path.dirname(__file__))
from openpyxl import load_workbook
from openpyxl.drawing.image import Image


def insertimg2excel(imgPath, ws, col, row):
    imgsize = (200, 200)  # 设置一个图像缩小的比例
    ws.column_dimensions[str(col)].width = imgsize[0] * 0.14  # 修改列A的宽

    img = Image(imgPath)  # 缩放图片
    img.width, img.height = imgsize

    ws.add_image(img, col + str(row))  # 图片 插入 A1 的位置上
    ws.row_dimensions[row].height = imgsize[1] * 0.78  # 修改列第1行的高

def insert_img():
    excel_path = os.path.join(basedir, "data.xlsx")
    wb = load_workbook(excel_path)
    ws = wb.active
    item_len = ws.max_row
    for i in range(1, item_len + 1):
        imgUrl = ws.cell(row=i+1, column=7).value
        if imgUrl:
            imgPath = './img/' + imgUrl.replace('http://vis.pku.edu.cn/', '').replace('/', '__')
            ws.cell(row=i+1, column=1, value='')
            try:
                insertimg2excel(imgPath, ws, 'A', i + 1)
            except:
                print("error:" + imgPath)
    wb.save(excel_path)  # 新的结果保存输出

def download_all_img():
    file_path = "./data.csv"
    csv_data = pd.read_csv(file_path, encoding='UTF-8', low_memory=False)  # 防止弹出警告
    df = pd.DataFrame(csv_data)
    labels=df['image_url']
    for cnt in range(0, len(labels)):
        img_url = df['image_url'][cnt]
        api_token = "fklasjfljasdlkfjlasjflasjfljhasdljflsdjflkjsadljfljsda"        
        try:
            download_img(img_url, api_token)
        except:
            print("error:" + img_url)

if __name__ == '__main__':
    # generate_labels()
    # download_all_img()
    insert_img()